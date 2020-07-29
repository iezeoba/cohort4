from openpyxl import load_workbook


def dictbuild():
    newlist = []
    customers = {}
    invoice = {}
    product = {}
    invoice_list = {}
    details = {}
    all_dicts = {}

    workbook = load_workbook(
        "C:/code/cohort4/python/excel_folder/exerciseexcel.xlsx")
    for table in workbook.sheetnames:
        newlist.append(table)
    for row in workbook[newlist[0]].iter_rows(min_row=2, max_row=workbook[newlist[0]].max_row, values_only=True):
        details = {'firstname': row[1], 'lastname': row[2],
                   'email': row[3], 'phone': row[4]}
        customers[row[0]] = details
    # print(customers)
    for row in workbook[newlist[1]].iter_rows(min_row=2, max_row=workbook[newlist[1]].max_row, values_only=True):
        details = {'customer_id': row[1], 'invoice_number': row[2],
                   'invoice_date': row[3], 'prepared_by': row[4]}
        invoice[row[0]] = details
    # print(invoice)
    for row in workbook[newlist[2]].iter_rows(min_row=2, max_row=workbook[newlist[2]].max_row, values_only=True):
        details = {'product_name': row[1], 'price': row[2]}
        product[row[0]] = details
    # print(product)
    for row in workbook[newlist[3]].iter_rows(min_row=2, max_row=workbook[newlist[3]].max_row, values_only=True):
        details = {'invoice_id': row[1],
                   'product_id': row[2], 'quantity': row[3]}
        invoice_list[row[0]] = details
    # print(invoice_list)

    all_dicts = {
        newlist[0]: customers,
        newlist[1]: invoice,
        newlist[2]: product,
        newlist[3]: invoice_list
    }
    return all_dicts


# print(dictbuild())
